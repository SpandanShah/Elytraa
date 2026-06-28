"""
predictor/views.py

API views for the University Predictor.

Endpoints:
    GET  /              → Render the main HTML page (index.html)
    POST /api/predict/  → Run prediction, return JSON results
    GET  /api/options/  → Return dropdown data (categories, boards, course keywords)
    POST /api/upload/   → (stub) Upload new Excel data file
"""

import logging
import io

from django.http import HttpResponse
from django.shortcuts import render
from django.views.decorators.csrf import ensure_csrf_cookie
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.exceptions import NotAuthenticated, AuthenticationFailed
from rest_framework.permissions import IsAuthenticated, AllowAny, IsAdminUser
from rest_framework.response import Response
from rest_framework.views import exception_handler as drf_exception_handler

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from accounts.models import PredictionLog

from .serializers import (
    DropdownOptionsSerializer,
    PredictionInputSerializer,
    PredictionResultSerializer,
)
from .services import PredictionService

logger = logging.getLogger(__name__)


def custom_exception_handler(exc, context):
    """Return {"success": false, "error": "Authentication required."} on 401/403."""
    response = drf_exception_handler(exc, context)
    if response is not None and isinstance(exc, (NotAuthenticated, AuthenticationFailed)):
        response.data = {"success": False, "error": "Authentication required."}
        response.status_code = status.HTTP_403_FORBIDDEN
    return response


@ensure_csrf_cookie
def index(request):
    """GET / — Renders the main HTML template."""
    return render(request, "predictor/index.html")


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def predict(request):
    """
    POST /api/predict/

    Request body (JSON):
        {
            "rank": 3000,
            "category": "GEN",
            "board": "GUJCET",
            "course_preferences": ["computer", "mechanical"],
            "min_results": 15
        }

    Response (JSON):
        {
            "success": true,
            "count": 28,
            "results": [ ... ],
            "error": null
        }
    """
    input_serializer = PredictionInputSerializer(data=request.data)

    if not input_serializer.is_valid():
        logger.warning("Invalid prediction input: %s", input_serializer.errors)
        return Response(
            {
                "success": False,
                "count": 0,
                "results": [],
                "error": input_serializer.errors,
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    data = input_serializer.validated_data

    # --- Access-tier checks (run before rate limiting) ---
    profile = request.user.profile

    # Task 3.1: Round access check
    requested_round = data.get("round")
    if requested_round == 2 and not profile.can_access_round2:
        return Response(
            {
                "success": False,
                "error": "Round 2 predictions are not available for free users.",
            },
            status=status.HTTP_403_FORBIDDEN,
        )
    if requested_round == 3 and not profile.can_access_round3:
        return Response(
            {
                "success": False,
                "error": "Round 3 predictions are not available for free users.",
            },
            status=status.HTTP_403_FORBIDDEN,
        )

    # Task 3.2: Board access check
    requested_board = data.get("board")
    if (
        requested_board
        and requested_board.upper() == "JEE"
        and not profile.can_access_jee
    ):
        return Response(
            {
                "success": False,
                "error": "JEE board predictions are not available for free users.",
            },
            status=status.HTTP_403_FORBIDDEN,
        )

    # Task 3.3: Daily rate-limit check
    today_count = PredictionLog.count_today(request.user)
    if today_count >= profile.daily_prediction_limit:
        return Response(
            {
                "success": False,
                "error": "Daily prediction limit reached. Please try again tomorrow.",
            },
            status=status.HTTP_429_TOO_MANY_REQUESTS,
        )

    # --- Run prediction service ---
    try:
        service = PredictionService()
        # Only pass inst_types filter for staff users (silently ignore otherwise)
        inst_types = None
        if request.user.is_staff:
            inst_types = data.get("inst_types")
        results = service.predict(
            student_rank=data["rank"],
            category=data.get("category"),
            board=data.get("board"),
            course_preferences=data.get("course_preferences"),
            min_results=data.get("min_results", 15),
            round_num=data.get("round"),
            inst_types=inst_types,
        )
    except ValueError as exc:
        logger.error("Prediction ValueError: %s", exc)
        return Response(
            {"success": False, "count": 0, "results": [], "error": str(exc)},
            status=status.HTTP_400_BAD_REQUEST,
        )
    except Exception as exc:
        logger.exception("Unexpected error during prediction: %s", exc)
        return Response(
            {
                "success": False,
                "count": 0,
                "results": [],
                "error": "Internal server error.",
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    # Task 3.3: Log successful prediction for rate limiting
    PredictionLog.objects.create(user=request.user)

    result_serializer = PredictionResultSerializer(results, many=True)
    return Response(
        {
            "success": True,
            "count": len(results),
            "results": result_serializer.data,
            "error": None,
        },
        status=status.HTTP_200_OK,
    )


@api_view(["GET"])
@permission_classes([AllowAny])
def options(request):
    """
    GET /api/options/

    Returns all dropdown options for the frontend:
        categories    — distinct reservation categories in the database
        boards        — distinct exam boards in the database
        course_keywords — all supported course preference keywords
    """
    try:
        data = PredictionService.get_available_options()
    except Exception as exc:
        logger.exception("Error fetching options: %s", exc)
        return Response(
            {"success": False, "data": None, "error": str(exc)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    serializer = DropdownOptionsSerializer(data)
    return Response(
        {"success": True, "data": serializer.data, "error": None},
        status=status.HTTP_200_OK,
    )


@api_view(["POST"])
def upload_data(request):
    """
    POST /api/upload/

    Stub — Excel upload via API is not yet implemented.
    Use the management command instead:
        python manage.py import_data
        python manage.py import_data --clear
    """
    return Response(
        {
            "success": False,
            "error": (
                "Not implemented. Run: "
                "python manage.py import_data  (or --clear for a fresh load)"
            ),
        },
        status=status.HTTP_501_NOT_IMPLEMENTED,
    )


@api_view(["POST"])
@permission_classes([IsAdminUser])
def export_predictions(request):
    """
    POST /api/export/

    Admin-only endpoint. Same input as /api/predict/ but returns an
    Excel file (.xlsx) instead of JSON. No rate limiting applied.
    """
    input_serializer = PredictionInputSerializer(data=request.data)

    if not input_serializer.is_valid():
        return Response(
            {"success": False, "error": input_serializer.errors},
            status=status.HTTP_400_BAD_REQUEST,
        )

    data = input_serializer.validated_data

    try:
        service = PredictionService()
        results = service.predict(
            student_rank=data["rank"],
            category=data.get("category"),
            board=data.get("board"),
            course_preferences=data.get("course_preferences"),
            min_results=data.get("min_results", 15),
            round_num=data.get("round"),
            inst_types=data.get("inst_types"),
        )
    except ValueError as exc:
        return Response(
            {"success": False, "error": str(exc)},
            status=status.HTTP_400_BAD_REQUEST,
        )
    except Exception as exc:
        logger.exception("Export error: %s", exc)
        return Response(
            {"success": False, "error": "Internal server error."},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    if not results:
        return Response(
            {"success": False, "error": "No results to export."},
            status=status.HTTP_404_NOT_FOUND,
        )

    # Build Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Predictions"

    # Header row
    headers = [
        "Institution", "Course", "Category", "Board",
        "Opening Rank", "Closing Rank", "Chance",
        "Course Preference", "University Score", "Round", "Institute Type",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="6366F1", end_color="6366F1", fill_type="solid"
    )
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, r in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=r["inst_name"])
        ws.cell(row=row_idx, column=2, value=r["course_name"])
        ws.cell(row=row_idx, column=3, value=r["category"])
        ws.cell(row=row_idx, column=4, value=r["board"])
        ws.cell(row=row_idx, column=5, value=r["opening_rank"])
        ws.cell(row=row_idx, column=6, value=r["closing_rank"])
        ws.cell(row=row_idx, column=7, value=r["chance"])
        ws.cell(row=row_idx, column=8, value=r["preferred_course"])
        ws.cell(row=row_idx, column=9, value=r["university_score"])
        ws.cell(row=row_idx, column=10, value=r["round"])
        ws.cell(row=row_idx, column=11, value=r.get("inst_type", ""))

    # Auto-fit column widths (approximate)
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

    # Write to bytes buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    rank = data["rank"]
    filename = f"Elytraa_Predictions_Rank_{rank}.xlsx"

    response = HttpResponse(
        buffer.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument"
            ".spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
